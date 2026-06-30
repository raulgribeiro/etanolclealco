"""
Atualizador do Dashboard Gestão Etanol — Clealco
Lê o Excel direto da pasta do OneDrive sincronizado no PC.
Sem login, sem internet, sem Azure.
"""

import json
import re
import os
import sys
import webbrowser
import subprocess
import shutil
import openpyxl

# ─────────────────────────────────────────────
#  CONFIGURAÇÕES
# ─────────────────────────────────────────────
# Caminho do Excel no OneDrive sincronizado
EXCEL_PATH = r"C:\Users\raulribeiro\OneDrive - CLEALCO AÇÚCAR E ÁLCOOL S.A\Compartilhados\Gestão Etanol-Dash.xlsx"

# Caminho do HTML (mesma pasta do script)
HTML_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gestao_etanol_dashboard.html")

# Pasta do repositório git (mesma pasta do script, se já for um repo clonado)
REPO_DIR = os.path.dirname(os.path.abspath(__file__))
# ─────────────────────────────────────────────


def publish_github():
    """Faz commit e push do HTML atualizado para o GitHub."""
    git = shutil.which("git")
    if not git:
        print("⚠ Git não encontrado — pulei a publicação online.")
        print("  Instale o Git (git-scm.com) para habilitar a publicação automática.")
        return

    # Garante que o HTML publicado se chama index.html (padrão do GitHub Pages)
    index_path = os.path.join(REPO_DIR, "index.html")
    if os.path.abspath(HTML_PATH) != os.path.abspath(index_path):
        shutil.copy(HTML_PATH, index_path)

    def run(args):
        return subprocess.run([git] + args, cwd=REPO_DIR,
                               capture_output=True, text=True, shell=False)

    # Verifica se já é um repositório git
    check = run(["status"])
    if check.returncode != 0:
        print("⚠ Esta pasta não é um repositório Git ainda.")
        print("  Configure o repositório uma vez seguindo as instruções no início do script.")
        return

    run(["add", "index.html"])
    commit = run(["commit", "-m", "Atualização automática dos dados"])
    if "nothing to commit" in (commit.stdout + commit.stderr):
        print("ℹ Nenhuma mudança nos dados desde a última publicação.")
        return

    push = run(["push"])
    output = push.stdout + push.stderr
    if push.returncode == 0:
        print("✔ Publicado no GitHub com sucesso!")
    else:
        print(f"⚠ Erro ao publicar:\n{output[-800:]}")




def find_excel():
    """Tenta encontrar o Excel em possíveis caminhos do OneDrive."""
    possiveis = [
        EXCEL_PATH,
        r"C:\Users\raulribeiro\OneDrive - Clealco\Documentos\Compartilhados\Gestão Etanol-Dash.xlsx",
        r"C:\Users\raulribeiro\OneDrive\Compartilhados\Gestão Etanol-Dash.xlsx",
        r"C:\Users\raulribeiro\OneDrive - Clealco\Gestão Etanol-Dash.xlsx",
    ]
    for path in possiveis:
        if os.path.exists(path):
            print(f"✔ Excel encontrado em:\n  {path}")
            return path

    # Se não encontrar, lista o que tem no OneDrive pra ajudar
    onedrive = rf"C:\Users\raulribeiro\OneDrive - Clealco"
    if os.path.exists(onedrive):
        print(f"\n⚠ Excel não encontrado. Conteúdo do OneDrive:")
        for root, dirs, files in os.walk(onedrive):
            level = root.replace(onedrive, "").count(os.sep)
            if level > 2:
                continue
            indent = "  " * level
            print(f"{indent}{os.path.basename(root)}/")
            for file in files:
                if file.endswith(".xlsx"):
                    print(f"{indent}  📄 {file}  ←  {os.path.join(root, file)}")
    raise Exception(
        "Excel não encontrado! Verifique se o OneDrive está sincronizado\n"
        "e cole o caminho correto no script (variável EXCEL_PATH)."
    )


def read_clm(wb):
    ws = wb["Fluxo CLM"]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        try:
            trader, corretora, distribuidora, contrato, volume, pedido, preco = (
                row[1], row[2], row[3], row[4], row[5], row[6], row[7]
            )
        except IndexError:
            continue
        if not volume and not contrato:
            continue
        rows.append({
            "trader":        str(trader or "").strip(),
            "corretora":     str(corretora or "").strip(),
            "distribuidora": str(distribuidora or "").strip(),
            "contrato":      str(contrato or "").strip(),
            "volume":        float(volume) if volume else 0,
            "pedido":        str(pedido or "").strip(),
            "preco":         float(preco) if preco else 0,
        })
    print(f"✔ Fluxo CLM: {len(rows)} registros")
    return rows


def read_qrz(wb):
    ws = wb["Fluxo QRZ"]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        try:
            safra, trader, corretora, distribuidora, contrato, volume, pedido, preco = (
                row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
            )
        except IndexError:
            continue
        if not volume and not contrato:
            continue
        rows.append({
            "safra":         str(safra or "").strip(),
            "trader":        str(trader or "").strip(),
            "corretora":     str(corretora or "").strip(),
            "distribuidora": str(distribuidora or "").strip(),
            "contrato":      str(contrato or "").strip(),
            "volume":        float(volume) if volume else 0,
            "pedido":        str(pedido or "").strip(),
            "preco":         float(preco) if preco else 0,
        })
    print(f"✔ Fluxo QRZ: {len(rows)} registros")
    return rows


def read_rbf(wb):
    ws = wb["Rel Base Fat"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            estab     = row[1]
            data_hora = row[5]
            cliente   = row[15]
            produto   = row[17]
            valor_nfe = row[21]
            qtd       = row[22]
            nat_op    = row[27]
            safra     = row[30]
        except IndexError:
            continue
        if not valor_nfe and not cliente:
            continue

        vlr_unit     = float(valor_nfe) / float(qtd) if valor_nfe and qtd and float(qtd) != 0 else 0
        vlr_unit_liq = (vlr_unit * 0.9043) - 20

        if hasattr(data_hora, "strftime"):
            data_str = data_hora.strftime("%Y-%m-%d")
        else:
            data_str = str(data_hora or "").strip()

        rows.append({
            "estab":         str(estab or "").strip(),
            "data":          data_str,
            "cliente":       str(cliente or "").strip(),
            "produto":       str(produto or "").strip(),
            "valor_nfe":     float(valor_nfe) if valor_nfe else 0,
            "qtd":           float(qtd) if qtd else 0,
            "nat_op":        str(nat_op or "").strip(),
            "vlr_unit":      round(vlr_unit, 4),
            "vlr_unit_liq":  round(vlr_unit_liq, 4),
            "safra":         str(safra or "").strip(),
        })
    print(f"✔ Rel Base Fat: {len(rows)} registros")
    return rows


def update_html(clm_data, qrz_data, rbf_data):
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    def replace_array(html, var_name, new_data):
        pattern = rf"(const {var_name} = )(\[.*?\]);"
        replacement = r"\g<1>" + json.dumps(new_data, ensure_ascii=False) + ";"
        new_html, n = re.subn(pattern, replacement, html, flags=re.DOTALL)
        if n == 0:
            raise Exception(f"Variável '{var_name}' não encontrada no HTML!")
        return new_html

    html = replace_array(html, "CLM_DATA", clm_data)
    html = replace_array(html, "QRZ_DATA", qrz_data)
    html = replace_array(html, "RBF_DATA", rbf_data)

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"✔ HTML atualizado!")


def main():
    print("=" * 52)
    print("   Atualizador Dashboard Gestão Etanol — Clealco")
    print("=" * 52)

    try:
        print("\n[1/4] Localizando Excel no OneDrive...")
        excel_path = find_excel()

        print("[2/4] Abrindo planilha...")
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

        print("[3/4] Lendo dados...")
        clm_data = read_clm(wb)
        qrz_data = read_qrz(wb)
        rbf_data = read_rbf(wb)
        wb.close()

        print("[4/4] Atualizando dashboard...")
        update_html(clm_data, qrz_data, rbf_data)

        print("\n✅ Dashboard atualizado! Publicando no GitHub...")
        publish_github()

        print("\n✅ Concluído! Abrindo dashboard no navegador...")
        webbrowser.open(f"file:///{HTML_PATH.replace(os.sep, '/')}")

    except Exception as e:
        print(f"\n❌ Erro: {e}")
        input("\nPressione Enter para fechar...")
        sys.exit(1)


if __name__ == "__main__":
    main()
