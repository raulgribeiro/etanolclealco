"""
Atualizador do Dashboard Gestão Etanol — Clealco
Lê o Excel direto da pasta do OneDrive sincronizado no PC.
Sem login, sem internet, sem Azure.
"""

import json
import re
import os
import sys
import time
import glob
import webbrowser
import subprocess
import shutil
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait

# ─────────────────────────────────────────────
#  CONFIGURAÇÕES
# ─────────────────────────────────────────────
# Caminho do Excel no OneDrive sincronizado
# Pasta do projeto (fixa)
PROJECT_DIR = r"C:\Users\raulribeiro\Documents\App Gestão Etanol"

# Caminho do Excel — baixado automaticamente via navegador (Selenium)
EXCEL_PATH = os.path.join(PROJECT_DIR, "Gestão Etanol-Dash.xlsx")

# Link do arquivo compartilhado (Marcello)
SHARE_LINK = "https://clealco-my.sharepoint.com/personal/marcello_pandin_clealco_com_br/Documents/Gestão Etanol.xlsx"

# O Chrome já baixa direto na pasta do projeto (sem precisar mover de Downloads)
DOWNLOADS_DIR = PROJECT_DIR

# Perfil do Chrome a reutilizar — para manter o login salvo entre execuções
CHROME_PROFILE_DIR = os.path.join(PROJECT_DIR, "_chrome_profile")

# Caminho do HTML (mesma pasta do script)
HTML_PATH = os.path.join(PROJECT_DIR, "index.html")

# Pasta do repositório git (mesma pasta do script, se já for um repo clonado)
REPO_DIR = PROJECT_DIR
# ─────────────────────────────────────────────


def publish_github():
    """Faz commit e push do HTML atualizado para o GitHub."""
    git = shutil.which("git")
    print(f"  → Git encontrado em: {git}")
    if not git:
        print("⚠ Git não encontrado — pulei a publicação online.")
        print("  Instale o Git (git-scm.com) para habilitar a publicação automática.")
        return

    print(f"  → Pasta do repositório: {REPO_DIR}")

    # Garante que o HTML publicado se chama index.html (padrão do GitHub Pages)
    index_path = os.path.join(REPO_DIR, "index.html")
    if os.path.abspath(HTML_PATH) != os.path.abspath(index_path):
        shutil.copy(HTML_PATH, index_path)

    def run(args):
        r = subprocess.run([git] + args, cwd=REPO_DIR,
                            capture_output=True, text=True, shell=False)
        print(f"  → git {' '.join(args)}")
        print(f"    stdout: {r.stdout.strip()[:300]}")
        print(f"    stderr: {r.stderr.strip()[:300]}")
        print(f"    returncode: {r.returncode}")
        return r

    # Verifica se já é um repositório git
    check = run(["status"])
    if check.returncode != 0:
        print("⚠ Esta pasta não é um repositório Git ainda.")
        print("  Configure o repositório uma vez seguindo as instruções no início do script.")
        return

    run(["add", "index.html"])
    commit = run(["commit", "-m", "Atualização automática dos dados"])
    if "nothing to commit" in (commit.stdout + commit.stderr):
        print("ℹ Nenhuma mudança nos dados desde a última publicação.")
        return

    push = run(["push"])
    output = push.stdout + push.stderr
    if push.returncode == 0:
        print("✔ Publicado no GitHub com sucesso!")
    else:
        print(f"⚠ Erro ao publicar:\n{output[-800:]}")




def check_recent_modification(path, minutos_aviso=10):
    """Avisa se o arquivo não foi modificado recentemente, sugerindo que o
    OneDrive pode ainda não ter sincronizado a última alteração."""
    import time
    mtime = os.path.getmtime(path)
    minutos_atras = (time.time() - mtime) / 60
    data_mod = __import__("datetime").datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M:%S")
    print(f"  → Última modificação do Excel: {data_mod}")

    if minutos_atras > minutos_avisar_threshold():
        print(f"\n⚠ ATENÇÃO: este arquivo não foi modificado nos últimos {int(minutos_atras)} minutos.")
        print("  Se você acabou de salvar uma alteração no SharePoint, o OneDrive")
        print("  pode ainda não ter sincronizado. Verifique se o ícone de nuvem")
        print("  do arquivo está com check verde ✅ antes de continuar.\n")
        resp = input("  Deseja continuar mesmo assim? (s/n): ").strip().lower()
        if resp != "s":
            print("\n❌ Operação cancelada pelo usuário. Aguarde a sincronização e tente de novo.")
            sys.exit(0)


def minutos_avisar_threshold():
    return 15  # ajuste aqui se quiser um limite diferente


def download_via_chrome():
    """Abre o Chrome, navega até o link do SharePoint e aguarda você baixar
    o arquivo manualmente (Arquivo → Salvar uma Cópia → Baixar uma Cópia).
    Detecta automaticamente quando o download termina."""

    os.makedirs(CHROME_PROFILE_DIR, exist_ok=True)

    # Limpa downloads antigos do mesmo nome para não confundir qual é o novo
    for old in glob.glob(os.path.join(DOWNLOADS_DIR, "Gestão Etanol*.xlsx")):
        try:
            os.remove(old)
        except OSError:
            pass

    options = Options()
    options.add_argument(f"--user-data-dir={CHROME_PROFILE_DIR}")
    options.add_argument("--profile-directory=Default")
    prefs = {
        "download.default_directory": DOWNLOADS_DIR,
        "download.prompt_for_download": False,
    }
    options.add_experimental_option("prefs", prefs)

    print("🌐 Abrindo o Chrome...")
    driver = webdriver.Chrome(options=options)

    try:
        driver.get(SHARE_LINK)
        print("✔ Página aberta.")
        print("\n  ➜ Faça login se for solicitado.")
        print("  ➜ Depois baixe o arquivo: Arquivo → Salvar uma Cópia → Baixar uma Cópia.")
        print("  ⏳ Aguardando o download terminar (até 3 minutos)...\n")

        downloaded = _wait_for_download(timeout=180)
        if not downloaded:
            raise Exception("Tempo esgotado esperando o download do Excel.")

        if os.path.abspath(downloaded) != os.path.abspath(EXCEL_PATH):
            if os.path.exists(EXCEL_PATH):
                os.remove(EXCEL_PATH)
            shutil.move(downloaded, EXCEL_PATH)
        print(f"✔ Excel baixado e salvo em:\n  {EXCEL_PATH}")

    finally:
        driver.quit()

    return EXCEL_PATH


def _wait_for_download(timeout=180):
    """Espera um arquivo .xlsx aparecer (e terminar de baixar) na pasta do projeto."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        candidatos = glob.glob(os.path.join(DOWNLOADS_DIR, "Gestão Etanol*.xlsx"))
        # Garante que não é um arquivo ainda baixando
        candidatos = [c for c in candidatos
                      if not c.endswith(".crdownload") and not c.endswith(".tmp")]
        if candidatos:
            time.sleep(1.5)  # garante que terminou de escrever no disco
            return max(candidatos, key=os.path.getmtime)
        time.sleep(1)
    return None


def find_excel():
    """Baixa o Excel automaticamente via Chrome (login manual só na 1ª vez)."""
    try:
        return download_via_chrome()
    except Exception as e:
        print(f"\n⚠ Não foi possível baixar automaticamente: {e}")
        # Fallback: procura arquivo já baixado manualmente antes
        pasta_projeto = PROJECT_DIR
        nomes_possiveis = ["Gestão Etanol-Dash.xlsx", "Gestão Etanol.xlsx"]
        possiveis = [EXCEL_PATH]
        for pasta in [pasta_projeto, DOWNLOADS_DIR]:
            for nome in nomes_possiveis:
                possiveis.append(os.path.join(pasta, nome))
        for path in possiveis:
            if os.path.exists(path):
                print(f"✔ Usando arquivo já existente em:\n  {path}")
                return path
        raise Exception(
            "Excel não encontrado nem foi possível baixar automaticamente.\n"
            "Baixe manualmente e salve na pasta do projeto como 'Gestão Etanol-Dash.xlsx'."
        )


def to_float(value):
    """Converte valores numéricos vindos do Excel, mesmo quando estão como
    texto mal formatado (ex: '3.163.88', '3.163,88', '3163,88', '1.234')."""
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return 0.0

    # Remove espaços e símbolos comuns
    s = s.replace(" ", "").replace("R$", "")

    # Caso tenha mais de um ponto (ex: '3.163.88' ou '3.163.880')
    if s.count(".") > 1:
        partes = s.split(".")
        # Junta tudo exceto a última parte (assume última = decimais)
        s = "".join(partes[:-1]) + "." + partes[-1]
    elif "," in s and "." in s:
        # Formato brasileiro: 3.163,88 → remove pontos de milhar, troca vírgula por ponto
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        # Só vírgula: assume separador decimal brasileiro
        s = s.replace(",", ".")

    try:
        return float(s)
    except ValueError:
        print(f"  ⚠ Valor não numérico ignorado: '{value}' → tratado como 0")
        return 0.0


def read_clm(wb):
    ws = wb["Fluxo CLM"]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        try:
            trader, corretora, distribuidora, contrato, volume, pedido, preco = (
                row[1], row[2], row[3], row[4], row[5], row[6], row[7]
            )
        except IndexError:
            continue
        if not volume and not contrato:
            continue
        rows.append({
            "trader":        str(trader or "").strip(),
            "corretora":     str(corretora or "").strip(),
            "distribuidora": str(distribuidora or "").strip(),
            "contrato":      str(contrato or "").strip(),
            "volume":        to_float(volume),
            "pedido":        str(pedido or "").strip(),
            "preco":         to_float(preco),
        })
    print(f"✔ Fluxo CLM: {len(rows)} registros")
    return rows


def read_qrz(wb):
    ws = wb["Fluxo QRZ"]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        try:
            safra, trader, corretora, distribuidora, contrato, volume, pedido, preco = (
                row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
            )
        except IndexError:
            continue
        if not volume and not contrato:
            continue
        rows.append({
            "safra":         str(safra or "").strip(),
            "trader":        str(trader or "").strip(),
            "corretora":     str(corretora or "").strip(),
            "distribuidora": str(distribuidora or "").strip(),
            "contrato":      str(contrato or "").strip(),
            "volume":        to_float(volume),
            "pedido":        str(pedido or "").strip(),
            "preco":         to_float(preco),
        })
    print(f"✔ Fluxo QRZ: {len(rows)} registros")
    return rows


def read_rbf(wb):
    ws = wb["Rel Base Fat"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            estab     = row[1]
            data_hora = row[5]
            cliente   = row[15]
            produto   = row[17]
            valor_nfe = row[21]
            qtd       = row[22]
            nat_op    = row[27]
            safra     = row[30]
        except IndexError:
            continue
        if not valor_nfe and not cliente:
            continue

        vlr_unit_qtd = to_float(qtd)
        vlr_unit     = to_float(valor_nfe) / vlr_unit_qtd if vlr_unit_qtd != 0 else 0
        vlr_unit_liq = (vlr_unit * 0.9043) - 20

        if hasattr(data_hora, "strftime"):
            data_str = data_hora.strftime("%Y-%m-%d")
        else:
            data_str = str(data_hora or "").strip()

        rows.append({
            "estab":         str(estab or "").strip(),
            "data":          data_str,
            "cliente":       str(cliente or "").strip(),
            "produto":       str(produto or "").strip(),
            "valor_nfe":     to_float(valor_nfe),
            "qtd":           to_float(qtd),
            "nat_op":        str(nat_op or "").strip(),
            "vlr_unit":      round(vlr_unit, 4),
            "vlr_unit_liq":  round(vlr_unit_liq, 4),
            "safra":         str(safra or "").strip(),
        })
    print(f"✔ Rel Base Fat: {len(rows)} registros")
    return rows


def read_exposicao(wb):
    """Lê a tabela resumo de Exposição na aba Caixa, linhas 5292-5298,
    colunas B (rótulo), C (Clementina), D (Queiroz), E (Total)."""
    ws = wb["Caixa"]
    rows = []
    for r in range(5293, 5299):  # 5292 é o cabeçalho, dados de 5293 a 5298
        natureza    = ws[f"B{r}"].value
        clementina  = ws[f"C{r}"].value
        queiroz     = ws[f"D{r}"].value
        total       = ws[f"E{r}"].value
        if not natureza:
            continue
        rows.append({
            "Natureza":   str(natureza).strip(),
            "Clementina": to_float(clementina),
            "Queiroz":    to_float(queiroz),
            "Total":      to_float(total),
        })
    print(f"✔ Exposição: {len(rows)} registros")
    return rows


def update_html(clm_data, qrz_data, rbf_data, exp_data):
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    def replace_array(html, var_name, new_data):
        pattern = rf"(const {var_name} = )(\[.*?\]);"
        replacement = r"\g<1>" + json.dumps(new_data, ensure_ascii=False) + ";"
        new_html, n = re.subn(pattern, replacement, html, flags=re.DOTALL)
        if n == 0:
            raise Exception(f"Variável '{var_name}' não encontrada no HTML!")
        return new_html

    html = replace_array(html, "CLM_DATA", clm_data)
    html = replace_array(html, "QRZ_DATA", qrz_data)
    html = replace_array(html, "RBF_DATA", rbf_data)
    html = replace_array(html, "EXP_DATA", exp_data)

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"✔ HTML atualizado!")


def main():
    print("=" * 52)
    print("   Atualizador Dashboard Gestão Etanol — Clealco")
    print("=" * 52)

    try:
        print("\n[1/4] Localizando Excel no OneDrive...")
        excel_path = find_excel()

        print("[2/4] Abrindo planilha...")
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

        print("[3/4] Lendo dados...")
        clm_data = read_clm(wb)
        qrz_data = read_qrz(wb)
        rbf_data = read_rbf(wb)
        exp_data = read_exposicao(wb)
        wb.close()

        print("[4/4] Atualizando dashboard...")
        update_html(clm_data, qrz_data, rbf_data, exp_data)

        print("\n✅ Dashboard atualizado! Publicando no GitHub...")
        publish_github()

        print("\n✅ Concluído! Abrindo dashboard no navegador...")
        webbrowser.open(f"file:///{HTML_PATH.replace(os.sep, '/')}")

    except Exception as e:
        print(f"\n❌ Erro: {e}")
        input("\nPressione Enter para fechar...")
        sys.exit(1)


if __name__ == "__main__":
    main()
